from datetime import datetime

from openpyxl import Workbook, load_workbook

wb = Workbook()

# grab the active worksheet
ws = wb.active

# Data can be assigned directly to cells
ws['A1'] = 42

# Rows can also be appended
ws.append([1, 2, 3])

# Python types will automatically be converted
ws['A2'] = datetime.now()

# Save the file
# wb.save("./resources/sample.xlsx")


wb2 = load_workbook("./resources/sample.xlsx")
print(wb2.sheetnames)
# ['Sheet']
