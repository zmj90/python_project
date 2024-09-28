from openpyxl import load_workbook
wb = load_workbook("sample.xlsx")
ws = wb["Sheet"]
for cells in ws["A1:C4"]:
    for cell in cells:
         print(cell.value, type(cell.value))
