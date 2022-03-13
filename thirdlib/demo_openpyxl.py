# from datetime import datetime
#
# from openpyxl import Workbook, load_workbook
#
# wb = Workbook()
#
# # grab the active worksheet
# ws = wb.active
#
# # Data can be assigned directly to cells
# ws['A1'] = 42
#
# # Rows can also be appended
# ws.append([1, 2, 3])
#
# # Python types will automatically be converted
# ws['A2'] = datetime.now()
#
# # Save the file
# # wb.save("./resources/sample.xlsx")
#
#
# wb2 = load_workbook("./resources/sample.xlsx")
# print(wb2.sheetnames)
# # ['Sheet']


# ---------------------------------------------------------------------
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

wb = Workbook()
dest_filename = './resources/empty_book.xlsx'
ws1 = wb.active
ws1.title = "range names"
for row in range(1, 40):
    ws1.append(range(600))
ws2 = wb.create_sheet(title="Pi")
ws2['F5'] = 3.14
ws3 = wb.create_sheet(title="Data")
for row in range(10, 20):
    for col in range(27, 54):
        _ = ws3.cell(column=col, row=row, value="{0}".format(get_column_letter(col)))
print(ws3['AA10'].value)
# AA
wb.save(filename=dest_filename)
